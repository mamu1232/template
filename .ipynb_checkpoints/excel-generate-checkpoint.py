#!/usr/bin/env python
# -*- coding: utf-8 -*-
import openpyxl as px
import pprint
import os
import shutil
#二次元配列をexcelに書き込む関数
def write_list_2d(sheet, l_2d, start_row, start_col):
    for y, row in enumerate(l_2d):
        for x, cell in enumerate(row):
            sheet.cell(row=start_row + y,
                       column=start_col + x,
                       value=l_2d[y][x])
#行数（target,condition,measure）の3つ
column_num=3
#新規作成または追加，ファイル名を取得
new_or_add      =input('new or add? n/a: ')
file_name       =input('file name:')
sheet_name      =input('sheet name:')
if new_or_add =='n':
    wb = px.Workbook()
    wb.save('{}.xlsx'.format(file_name))
elif new_or_add =='a':
    wb = px.load_workbook('{}.xlsx'.format(file_name))
ws = wb.create_sheet(sheet_name)
sheet = wb[sheet_name]

## 列名まで含めたフォント設定がわからない
# font = px.styles.Font(
#         name = "游ゴシック",
#         size = 11,
#     )
# for row_num in range(1,50):
#     for column_num in range(1,50):
#         style = sheet.cell(row=row_num, column=column_num)
#         style.font = font
##

#試料・条件・測定値の数を取得
target_num      =int(input('How many targets: '))
condition_num   =int(input('How many conditions: '))
measure_num     =int(input('How many measurements: '))
#試料・条件・測定値のリストを取得
target_list     =[]
target_valist   =[]
condition_list  =[]
condition_valist=[]
measure_list    =[]
unit_list       =[]
for num in range(0,target_num):
    target    =input('target{}: '.format(num+1))
    target_list.append(target)
    target_val=input('target{}_val: '.format(num+1))
    target_valist.append(target_val)
for num in range(0,condition_num):
    condition =input('condition{}: '.format(num+1))
    condition_list.append(condition)
    condition_val =input('condition{}_val: '.format(num+1))
    condition_valist.append(condition_val)
for num in range(0,measure_num):
    measure   =input('measure{}: '.format(num+1))
    measure_list.append(measure)
    unit      = input('unit of measure{}:'.format(num+1))
    unit_list.append(unit)
#excelに書き込むに次元配列の作成
raw_num = 1
for num in (target_num,condition_num,measure_num):
    raw_num = raw_num*num
excel=[[None]*raw_num for i in range(column_num)]
#行ごとに代入
for i in range(0,target_num):
    excel[0][i*condition_num*measure_num] = target_list[i]
    for j in range(0,condition_num):
        excel[1][j*measure_num + i*condition_num*measure_num] = condition_list[j]
        for k in range(0,measure_num):
            excel[2][k+j*measure_num + i*condition_num*measure_num] = measure_list[k]+'_{}{}'.format(target_valist[i],condition_valist[j])+'/{}'.format(unit_list[k])
#書込みと保存
write_list_2d(sheet,excel,1,1)
wb.save('{}.xlsx'.format(file_name))